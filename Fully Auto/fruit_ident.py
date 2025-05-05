# fruit_ident.py

import numpy as np
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing import image
import openpyxl
from datetime import datetime

# === Load Model Only Once ===
model_path = r"C:\Users\hitis\Desktop\Applied Robotics\Report\Codes\Fully Auto\best_model.keras"
model = load_model(model_path)

# === Class Labels ===
class_labels = [
    'banana_ripe', 'banana_rotten', 'banana_unripe',
    'strawberry_ripe', 'strawberry_rotten', 'strawberry_unripe',
    'tomato_ripe', 'tomato_rotten', 'tomato_unripe'
]

def classify_and_log(img_path):
    # === Preprocess Image ===
    img = image.load_img(img_path, target_size=(224, 224))
    img_array = image.img_to_array(img) / 255.0
    img_array = np.expand_dims(img_array, axis=0)

    # === Predict ===
    predictions = model.predict(img_array)
    class_index = np.argmax(predictions)
    predicted_class = class_labels[class_index]
    fruit, quality = predicted_class.split('_')

    # === Log to Excel ===
    excel_path = r"C:\Users\hitis\Desktop\Applied Robotics\Report\Codes\Fully Auto\Data logger.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    next_row = sheet.max_row + 1
    now = datetime.now()
    sheet.cell(row=next_row, column=1).value = fruit
    sheet.cell(row=next_row, column=2).value = quality
    sheet.cell(row=next_row, column=3).value = now.strftime("%Y-%m-%d")
    sheet.cell(row=next_row, column=4).value = now.strftime("%H:%M:%S")
    wb.save(excel_path)

    print(f"Logged: {fruit}, {quality}, {now.strftime('%Y-%m-%d')}, {now.strftime('%H:%M:%S')}")
    return fruit, quality
